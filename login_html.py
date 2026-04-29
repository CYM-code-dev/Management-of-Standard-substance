# login_html.py - 完整整合版（修复真实姓名显示）
import sys
import os
import json
import time
import threading
import datetime
import base64
import hashlib
import re
from io import BytesIO
from urllib.parse import urlencode

import requests
import urllib3
from flask import Flask, session, jsonify, request, redirect, render_template_string, render_template, send_from_directory
from flask_cors import CORS
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment
import xlrd
import xlwt
from xlutils.copy import copy as xl_copy
from xlutils.styles import Styles

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)
app.secret_key = 'a-fixed-secret-key-for-login-system-2026'
CORS(app, supports_credentials=True)

# ==================== 配置文件路径 ====================
CONFIG_FILE = 'config.json'


def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'excelPath': '', 'certPath': ''}


def save_config(excel_path, cert_path):
    config = {'excelPath': excel_path, 'certPath': cert_path}
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2)


# ==================== 远程系统连接类 ====================
def md5_1024_times(text: str) -> str:
    current = text.encode('utf-8')
    for _ in range(1024):
        current = hashlib.md5(current).hexdigest().encode('utf-8')
    return current.decode()


class RemoteSystem:
    def __init__(self, sess_id, user_agent=None):
        self.sess_id = sess_id
        self.session = requests.Session()
        self.base_url = "http://192.168.12.234:60015"
        self.session.verify = False
        self.current_user = None
        self.current_pid = None
        self.current_real_name = None
        self.captcha_session = None
        self.keep_alive_flag = False
        self.keep_alive_thread = None
        self.keep_alive_interval = 3600
        self.user_agent = user_agent
        self.update_headers()

    def set_user_agent(self, ua_string):
        self.user_agent = ua_string
        self.update_headers()

    def update_headers(self):
        default_ua = "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36"
        self.headers = {
            "User-Agent": self.user_agent or default_ua,
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Origin": self.base_url,
            "Referer": f"{self.base_url}/web/login.html",
            "Connection": "keep-alive",
            "Host": "192.168.12.234:60015"
        }
        self.session.headers.update(self.headers)

    def get_captcha_image(self):
        try:
            temp_session = requests.Session()
            temp_session.verify = False
            temp_session.headers.update(self.headers)
            ts = str(int(time.time() * 1000))
            url = f"{self.base_url}/detectionManager/core/security/validatecodes?{ts}&r={ts}"
            resp = temp_session.get(url)
            if resp.status_code == 200:
                self.captcha_session = temp_session
                return Image.open(BytesIO(resp.content))
            return None
        except Exception as e:
            print(f"获取验证码失败: {e}")
            return None

    def login(self, username, plain_password, captcha):
        password_hash = md5_1024_times(plain_password)
        login_data = {"account": username, "password": password_hash, "validCode": captcha}
        try:
            if self.captcha_session:
                resp = self.captcha_session.post(f"{self.base_url}/detectionManager/core/security/login", data=login_data)
                if resp.status_code == 200:
                    for cookie in self.captcha_session.cookies:
                        self.session.cookies.set(cookie.name, cookie.value)
            else:
                resp = self.session.post(f"{self.base_url}/detectionManager/core/security/login", data=login_data)
            if resp.status_code != 200:
                return False, f"请求失败，状态码: {resp.status_code}"
            result = resp.json()
            if result.get("success"):
                self.current_user = username
                # 获取真实姓名
                real_name = self._fetch_user_info()
                if not real_name:
                    real_name = result.get("resultData", {}).get("nickName", username)
                self.current_real_name = real_name
                self._save_session()
                self.start_keep_alive()
                return True, real_name
            else:
                error_msg = result.get("errorCtx", {}).get("errorMsg", "登录失败")
                return False, "验证码错误" if "验证码" in error_msg else error_msg
        except Exception as e:
            return False, f"登录异常: {str(e)}"

    def _fetch_user_info(self):
        """获取真实姓名，优先从 getLoginUser 接口获取 realName，其次从 users/info 接口获取"""
        try:
            # 接口1
            resp = self.session.get(f"{self.base_url}/detectionManager/core/security/getLoginUser")
            if resp.status_code == 200:
                data = resp.json()
                if data.get("success"):
                    user_info = data.get("resultData", {}).get("userInfo", {})
                    pid = user_info.get("id")
                    real_name = user_info.get("realName")
                    if pid:
                        self.current_pid = str(pid)
                    if real_name:
                        return real_name
            # 接口2
            resp2 = self.session.get(f"{self.base_url}/detectionManager/core/users/info")
            if resp2.status_code == 200:
                data2 = resp2.json()
                if data2.get("success"):
                    user_info = data2.get("resultData", {}).get("userInfo", {})
                    pid = user_info.get("id")
                    real_name = user_info.get("realName")
                    if pid:
                        self.current_pid = str(pid)
                    if real_name:
                        return real_name
        except Exception as e:
            print(f"获取用户信息异常: {e}")
        return None

    def _save_session(self):
        sess_file = f"session_{self.current_user}.json"
        cookies_dict = {c.name: c.value for c in self.session.cookies}
        info = {
            "username": self.current_user, "pid": self.current_pid, "real_name": self.current_real_name,
            "login_time": time.strftime("%Y-%m-%d %H:%M:%S"),
            "cookies": cookies_dict, "headers": dict(self.session.headers)
        }
        with open(sess_file, "w", encoding="utf-8") as f:
            json.dump(info, f, indent=2)

    def load_session(self):
        if not self.current_user: return False
        sess_file = f"session_{self.current_user}.json"
        if not os.path.exists(sess_file): return False
        try:
            with open(sess_file, 'r', encoding='utf-8') as f:
                info = json.load(f)
            if time.time() - time.mktime(time.strptime(info["login_time"], "%Y-%m-%d %H:%M:%S")) > 24*3600:
                os.remove(sess_file)
                return False
            for name, value in info["cookies"].items():
                self.session.cookies.set(name, value)
            self.session.headers.update(info["headers"])
            self.current_user = info["username"]
            self.current_pid = info.get("pid")
            self.current_real_name = info.get("real_name", info["username"])
            return True
        except:
            return False

    def verify_session(self):
        if not self.current_user: return False
        try:
            resp = self.session.get(f"{self.base_url}/detectionManager/core/security/getLoginUser")
            return resp.status_code == 200 and resp.json().get("success")
        except:
            return False

    def logout(self):
        self.stop_keep_alive()
        if self.current_user:
            sess_file = f"session_{self.current_user}.json"
            if os.path.exists(sess_file): os.remove(sess_file)
        self.current_user = self.current_pid = self.current_real_name = None
        self.session.cookies.clear()

    def start_keep_alive(self):
        if not self.should_keep_alive() or self.keep_alive_flag: return
        self.keep_alive_flag = True
        self.keep_alive_thread = threading.Thread(target=self._keep_alive_worker, daemon=True)
        self.keep_alive_thread.start()

    def stop_keep_alive(self):
        self.keep_alive_flag = False
        if self.keep_alive_thread: self.keep_alive_thread.join(timeout=1)

    def should_keep_alive(self):
        return datetime.datetime.now().hour < 20

    def _keep_alive_worker(self):
        while self.keep_alive_flag and self.should_keep_alive() and self.current_user:
            try:
                self.session.get(f"{self.base_url}/detectionManager/core/security/getLoginUser")
            except: pass
            for _ in range(self.keep_alive_interval):
                if not self.keep_alive_flag: break
                time.sleep(1)


user_systems = {}
def get_system():
    sess_id = session.get('sess_id')
    if not sess_id:
        sess_id = os.urandom(16).hex()
        session['sess_id'] = sess_id
    if sess_id not in user_systems:
        user_systems[sess_id] = RemoteSystem(sess_id)
    return user_systems[sess_id]

# ==================== 路由 ====================
@app.route('/')
def index():
    if not session.get('logged_in'): return redirect('/login')
    return redirect('/organic-std')

@app.route('/login')
def login_page():
    error = request.args.get('error', '')
    username = request.args.get('username', '')
    system = get_system()
    img = system.get_captcha_image()
    captcha_url = ''
    if img:
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        captcha_url = f"data:image/png;base64,{base64.b64encode(buffered.getvalue()).decode()}"
    return render_template('login.html', error=error, error_json=json.dumps(error), username=username,
                           captcha_url=captcha_url)

@app.route('/organic-std')
def organic_std_page():
    return render_template('OrganicStd.html')

@app.route('/api/captcha')
def get_captcha():
    system = get_system()
    img = system.get_captcha_image()
    if img:
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        return jsonify({"success": True, "image": f"data:image/png;base64,{base64.b64encode(buffered.getvalue()).decode()}"})
    return jsonify({"success": False, "message": "获取验证码失败"})

@app.route('/api/login', methods=['POST'])
def login():
    if request.is_json:
        data = request.json
        username = data.get('username'); plain_password = data.get('password'); captcha = data.get('captcha'); client_ua = data.get('client_ua'); is_json = True
    else:
        username = request.form.get('username'); plain_password = request.form.get('password'); captcha = request.form.get('captcha'); client_ua = request.form.get('client_ua'); is_json = False

    # 如果是表单提交且 session 已登录，直接重定向，不再校验
    if not is_json and session.get('logged_in'):
        next_page = request.form.get('next', '/')
        return redirect(next_page)

    if not username or not plain_password or not captcha:
        if is_json: return jsonify({"success": False, "message": "账号、密码和验证码不能为空"})
        else: return redirect(f'/login?error=请填写完整信息&username={username}')

    system = get_system()
    if client_ua: system.set_user_agent(client_ua)

    success, display_name = system.login(username, plain_password, captcha)

    if success:
        session['logged_in'] = True
        session['username'] = username
        session['display_name'] = display_name
        session['pid'] = system.current_pid
        if is_json: return jsonify({"success": True, "message": "登录成功", "display_name": display_name})
        else: return redirect(request.form.get('next', '/'))
    else:
        if is_json: return jsonify({"success": False, "message": display_name})
        else: return redirect(f'/login?error={display_name}&username={username}')
@app.route('/api/status')
def status():
    if session.get('logged_in'):
        return jsonify({"logged_in": True, "username": session.get('username'), "display_name": session.get('display_name')})
    system = get_system()
    if system.current_user and system.verify_session():
        session['logged_in'] = True
        session['username'] = system.current_user
        session['display_name'] = system.current_real_name or system.current_user
        session['pid'] = system.current_pid
        return jsonify({"logged_in": True, "username": system.current_user, "display_name": session['display_name']})
    return jsonify({"logged_in": False})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    get_system().logout()
    return jsonify({"success": True})

@app.route('/api/query')
def query():
    if not session.get('logged_in'): return jsonify({"success": False, "message": "未登录"})
    pid = session.get('pid')
    username = session.get('username')
    if not pid:
        system = get_system()
        if system.current_pid: pid = session['pid'] = system.current_pid
        else: return jsonify({"success": False, "message": "无法获取用户PID"})
    keyword = request.args.get('keyword', '').strip()
    if not keyword: return jsonify({"success": False, "message": "请输入查询关键词"})
    system = get_system()
    if system.current_user != username:
        system.current_user = username
        system.load_session()
    params = {
        "_search": "false", "nd": str(int(time.time()*1000)), "pageSize": 30, "pageNo": 1, "sidx": "", "sord": "asc",
        "type": "CONSUMABLE_DIR_TYPE_STANDARD_SUBSTANCE", "orgName": "", "groupId": "", "status": "normal",
        "keyword": keyword, "state": "normal", "pid": pid, "pname": username, "loginId": pid
    }
    try:
        resp = system.session.get(f"{system.base_url}/detectionManager/manager/consumableBill/pageObj", params=params)
        if resp.status_code != 200:
            if resp.status_code in (401,403): session.pop('logged_in', None); system.logout(); return jsonify({"success": False, "message": "远程会话已失效"})
            return jsonify({"success": False, "message": f"请求失败，状态码: {resp.status_code}"})
        result = resp.json()
        if result.get("success"): return jsonify({"success": True, "data": result.get("resultData", {}).get("voList", [])})
        error_msg = result.get("errorCtx", {}).get("errorMsg", "查询失败")
        if "未登录" in error_msg or "login" in error_msg.lower(): session.pop('logged_in', None); system.logout(); return jsonify({"success": False, "message": "远程会话已失效"})
        return jsonify({"success": False, "message": error_msg})
    except Exception as e:
        return jsonify({"success": False, "message": f"查询异常: {str(e)}"})

@app.route('/api/config', methods=['GET', 'POST'])
def handle_config():
    if request.method == 'GET': return jsonify(load_config())
    data = request.get_json()
    excel_path = data.get('excelPath', '').strip()
    cert_path = data.get('certPath', '').strip()
    save_config(excel_path, cert_path)
    excel_ok = not excel_path or os.path.exists(os.path.dirname(excel_path))
    cert_ok = not cert_path or os.path.exists(cert_path)
    if excel_ok and cert_ok: return jsonify({"success": True})
    msg = []
    if not excel_ok: msg.append("Excel 路径不可访问")
    if not cert_ok: msg.append("证书路径不可访问")
    return jsonify({"success": False, "message": "；".join(msg)}), 400

@app.route('/organic_excel/<path:filename>')
def serve_excel(filename):
    config = load_config()
    excel_dir = os.path.dirname(config.get('excelPath', ''))
    if not excel_dir:
        return jsonify({"success": False, "message": "未配置 Excel 路径"}), 400
    return send_from_directory(excel_dir, filename)

@app.route('/certificates/<path:filename>')
def serve_cert(filename):
    config = load_config()
    cert_dir = config.get('certPath', '')
    if not cert_dir:
        return jsonify({"success": False, "message": "未配置证书路径"}), 400
    return send_from_directory(cert_dir, filename)


# ==================== 智能插入写入 Excel ====================
def get_prefix_type(original_id):
    if not original_id: return 'number'
    first = original_id[0].upper()
    if first == 'D': return 'D'
    if first == 'E': return 'E'
    return 'number'

def extract_number(original_id):
    if not original_id: return 0
    m = re.search(r'\d+', original_id)
    return int(m.group()) if m else 0

def is_row_empty(row_values):
    return all(v is None or str(v).strip() == '' for v in row_values)

def parse_existing_records_xlsx(ws):
    records = []
    for row in ws.iter_rows(min_row=2):
        row_vals = [c.value for c in row]
        if is_row_empty(row_vals): continue
        original_id = str(row_vals[0]) if row_vals[0] is not None else ''
        lab_no = str(row_vals[2]) if len(row_vals) > 2 and row_vals[2] is not None else ''
        records.append({'row_index': row[0].row, 'original_id': original_id, 'lab_no': lab_no})
    return records

def parse_existing_records_xls(sheet):
    records = []
    for row_idx in range(1, sheet.nrows):
        row_vals = [sheet.cell_value(row_idx, c) for c in range(sheet.ncols)]
        if is_row_empty(row_vals): continue
        original_id = str(row_vals[0]) if row_vals[0] else ''
        lab_no = str(row_vals[2]) if len(row_vals) > 2 and row_vals[2] else ''
        records.append({'row_index': row_idx, 'original_id': original_id, 'lab_no': lab_no})
    return records

def find_insert_position(records, new_id):
    new_type = get_prefix_type(new_id)
    new_num = extract_number(new_id)
    numbers = [r for r in records if get_prefix_type(r['original_id']) == 'number']
    d_records = [r for r in records if get_prefix_type(r['original_id']) == 'D']
    e_records = [r for r in records if get_prefix_type(r['original_id']) == 'E']
    numbers.sort(key=lambda x: extract_number(x['original_id']))
    d_records.sort(key=lambda x: extract_number(x['original_id']))
    e_records.sort(key=lambda x: extract_number(x['original_id']))
    if new_type == 'number':
        idx = 0
        for r in numbers:
            if extract_number(r['original_id']) < new_num: idx += 1
            else: break
        return 2 + idx
    elif new_type == 'D':
        base = 2 + len(numbers)
        idx = 0
        for r in d_records:
            if extract_number(r['original_id']) < new_num: idx += 1
            else: break
        return base + idx
    else:
        base = 2 + len(numbers) + len(d_records)
        idx = 0
        for r in e_records:
            if extract_number(r['original_id']) < new_num: idx += 1
            else: break
        return base + idx

def check_duplicate_labno(records, new_lab_no):
    if not new_lab_no or new_lab_no.strip() == '/' or new_lab_no.strip() == '':
        return False
    return any(r['lab_no'] == new_lab_no for r in records)


def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = src.font.copy()
        dst.border = src.border.copy()
        dst.fill = src.fill.copy()
        dst.number_format = src.number_format
        dst.protection = src.protection.copy()
        dst.alignment = src.alignment.copy()


@app.route('/api/add_to_excel', methods=['POST'])
def add_to_excel():
    if not session.get('logged_in'):
        return jsonify({"success": False, "message": "未登录"}), 401
    try:
        data = request.get_json()
        record = data.get('record')
        if not record: return jsonify({"success": False, "message": "无数据"}), 400
        config = load_config()
        excel_path = config.get('excelPath', '').strip()
        if not excel_path: return jsonify({"success": False, "message": "未配置 Excel 路径"}), 400
        ext = os.path.splitext(excel_path)[1].lower()
        if ext not in ['.xls', '.xlsx']: return jsonify({"success": False, "message": "仅支持 .xls 或 .xlsx"}), 400

        sheet_name = '有机标准物质'
        headers = ['编号', '组别', '实验室编号', '标品名称', 'CAS号', '规格/浓度', '生产商',
                   '有效期', '存放地点', '入库日期', '使用情况', '备注']
        new_lab_no = record.get('labNo', '')
        new_original_id = record.get('originalId', '')
        new_row_data = [new_original_id, record.get('group', ''), new_lab_no, record.get('name', ''),
                        record.get('cas', ''), record.get('spec', ''), record.get('manufacturer', ''),
                        record.get('expiry', ''), record.get('location', ''), record.get('storageDate', ''),
                        record.get('usage', ''), record.get('remarks', '')]

        if ext == '.xlsx':
            wb = openpyxl.load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal='center')
            else:
                ws = wb[sheet_name]
            records = parse_existing_records_xlsx(ws)
            if check_duplicate_labno(records, new_lab_no):
                wb.close()
                return jsonify({"success": False, "message": f"实验室编号 {new_lab_no} 已存在"}), 400
            insert_row = find_insert_position(records, new_original_id)
            ws.insert_rows(insert_row)
            for col, val in enumerate(new_row_data, 1):
                ws.cell(row=insert_row, column=col, value=val)

            # 复制样式
            source_row = insert_row - 1
            while source_row >= 1:
                row_empty = all(ws.cell(row=source_row, column=c).value is None for c in range(1, len(headers)+1))
                if not row_empty: break
                source_row -= 1
            if source_row < 1:
                source_row = insert_row + 1
                max_row = ws.max_row
                while source_row <= max_row:
                    row_empty = all(ws.cell(row=source_row, column=c).value is None for c in range(1, len(headers)+1))
                    if not row_empty: break
                    source_row += 1
            if 1 <= source_row <= ws.max_row:
                for col in range(1, len(new_row_data) + 1):
                    src_cell = ws.cell(row=source_row, column=col)
                    dst_cell = ws.cell(row=insert_row, column=col)
                    copy_cell_style(src_cell, dst_cell)

            wb.save(excel_path)
            wb.close()

        else:  # .xls
            rb = xlrd.open_workbook(excel_path, formatting_info=True)
            wb = xl_copy(rb)
            styles = Styles(rb)
            if sheet_name in rb.sheet_names():
                sheet_rb = rb.sheet_by_name(sheet_name)
                ws = wb.get_sheet(sheet_name)
                records = parse_existing_records_xls(sheet_rb)
                if check_duplicate_labno(records, new_lab_no):
                    return jsonify({"success": False, "message": f"实验室编号 {new_lab_no} 已存在"}), 400
                insert_row_idx = find_insert_position(records, new_original_id) - 1
                for r in range(sheet_rb.nrows - 1, insert_row_idx - 1, -1):
                    for c in range(sheet_rb.ncols):
                        val = sheet_rb.cell_value(r, c)
                        try:
                            style = styles[r][c]
                        except:
                            style = xlwt.easyxf()
                        ws.write(r + 1, c, val, style)
                for c, val in enumerate(new_row_data):
                    try:
                        style = styles[insert_row_idx - 1][c] if insert_row_idx > 0 else styles[0][c]
                    except:
                        style = xlwt.easyxf()
                    ws.write(insert_row_idx, c, val, style)
            else:
                ws = wb.add_sheet(sheet_name)
                for c, h in enumerate(headers): ws.write(0, c, h)
                for c, val in enumerate(new_row_data): ws.write(1, c, val)
            wb.save(excel_path)

        return jsonify({"success": True, "message": "数据已同步到 Excel"})
    except Exception as e:
        return jsonify({"success": False, "message": f"写入 Excel 失败: {str(e)}"}), 500


@app.route('/api/update_lims_unit', methods=['POST'])
def update_lims_unit():
    if not session.get('logged_in'):
        return jsonify({"success": False, "message": "未登录"}), 401
    try:
        data = request.get_json()
        record_id = data.get('recordId')
        new_unit = data.get('concentrationUnitName', '').strip()
        lims_item = data.get('limsItem')
        if not record_id or not lims_item:
            return jsonify({"success": False, "message": "缺少记录ID或原始数据"}), 400

        system = get_system()
        if not system.current_user:
            return jsonify({"success": False, "message": "远程会话已失效，请重新登录"})

        form_data = {}
        form_data['concentrationUnitName'] = new_unit
        form_data['pid'] = system.current_pid
        form_data['pname'] = system.current_user
        form_data['loginId'] = system.current_pid
        form_data['_method'] = 'PUT'

        url = f"{system.base_url}/detectionManager/manager/consumableBill/{record_id}"
        for k in list(form_data.keys()):
            if form_data[k] is None:
                form_data[k] = ''
        resp = system.session.post(url, data=form_data)
        if resp.status_code != 200:
            try: detail = resp.text[:2000]
            except: pass
            print(f"[LIMS UPDATE] response: {detail}")
            return jsonify({"success": False, "message": f"LIMS 请求失败，状态码: {resp.status_code}，{detail}"})
        result = resp.json()
        if result.get('success'):
            return jsonify({"success": True, "message": "浓度单位已同步到 LIMS"})
        else:
            return jsonify({"success": False, "message": result.get('errorDesc', '更新失败')})
    except Exception as e:
        return jsonify({"success": False, "message": f"LIMS 同步异常: {str(e)}"}), 500


if __name__ == '__main__':
    if not os.path.exists('templates'): os.makedirs('templates')
    print("Flask 服务启动，访问以下地址：")
    print("  登录页面: http://127.0.0.1:5000/login")
    print("  耗材查询主页: http://127.0.0.1:5000/")
    print("  有机标准品管理: http://127.0.0.1:5000/organic-std")
    app.run(host='0.0.0.0', port=5000, debug=True)